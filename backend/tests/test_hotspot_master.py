from __future__ import annotations

import warnings
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import pytest
import shapefile
from openpyxl import load_workbook
from shapely import make_valid
from shapely.geometry import Point, shape

from app.config import POLYGON_SHADOW_GEOMETRY_VERSION
from app.ingest.hotspot_master import (
    EXPECTED_HOTSPOT_COUNT,
    SEOUL_WGS84_BBOX,
    HotspotMasterError,
    load_hotspot_geometry_master,
    load_hotspot_master,
)


FIXTURES = Path(__file__).resolve().parents[1] / "fixtures"
XLSX = FIXTURES / "seoul_hotspots_master.xlsx"
SHAPEFILE_ZIP = FIXTURES / "seoul_hotspot_areas.zip"


def _load_fixture_workbook():
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="^Unknown extension is not supported and will be removed$",
            category=UserWarning,
        )
        return load_workbook(XLSX)


def _write_modified_shapefile_zip(
    target: Path,
    *,
    omit_last_record: bool = False,
    duplicate_second_code: bool = False,
    replacement_prj: bytes | None = None,
) -> None:
    """Create a deterministic corrupt derivative of the real binary fixture."""

    with ZipFile(SHAPEFILE_ZIP) as source_zip:
        names = source_zip.namelist()
        stem = next(name[:-4] for name in names if name.lower().endswith(".shp"))
        reader = shapefile.Reader(
            shp=BytesIO(source_zip.read(stem + ".shp")),
            shx=BytesIO(source_zip.read(stem + ".shx")),
            dbf=BytesIO(source_zip.read(stem + ".dbf")),
            encoding="utf-8",
        )
        records = list(reader.iterShapeRecords())
        fields = reader.fields[1:]
        shape_type = reader.shapeType
        reader.close()

        if omit_last_record:
            records = records[:-1]

        shp_bytes = BytesIO()
        shx_bytes = BytesIO()
        dbf_bytes = BytesIO()
        writer = shapefile.Writer(
            shp=shp_bytes,
            shx=shx_bytes,
            dbf=dbf_bytes,
            shapeType=shape_type,
            encoding="utf-8",
        )
        for field in fields:
            writer.field(*field)
        area_cd_index = next(
            index for index, field in enumerate(fields) if field[0] == "AREA_CD"
        )
        first_code = records[0].record[area_cd_index]
        for index, shape_record in enumerate(records):
            values = list(shape_record.record)
            if duplicate_second_code and index == 1:
                values[area_cd_index] = first_code
            writer.shape(shape_record.shape)
            writer.record(*values)
        writer.close()

        replacements = {
            stem + ".shp": shp_bytes.getvalue(),
            stem + ".shx": shx_bytes.getvalue(),
            stem + ".dbf": dbf_bytes.getvalue(),
        }
        if replacement_prj is not None:
            replacements[stem + ".prj"] = replacement_prj

        with ZipFile(target, "w") as target_zip:
            for name in names:
                if name.endswith("/"):
                    continue
                target_zip.writestr(name, replacements.get(name, source_zip.read(name)))


def test_loads_and_deterministically_joins_official_binary_fixtures() -> None:
    records = load_hotspot_master(XLSX, SHAPEFILE_ZIP)

    assert len(records) == EXPECTED_HOTSPOT_COUNT
    assert [record.area_cd for record in records] == sorted(
        record.area_cd for record in records
    )
    assert len({record.area_cd for record in records}) == EXPECTED_HOTSPOT_COUNT

    by_code = {record.area_cd: record for record in records}
    assert by_code["POI001"].name == "강남 MICE 관광특구"
    assert by_code["POI001"].category == "관광특구"
    assert by_code["POI007"].name == "홍대 관광특구"
    assert by_code["POI131"].name == "숭례문"


def test_geometry_loader_exposes_normalized_versioned_official_polygons() -> None:
    records = load_hotspot_geometry_master(XLSX, SHAPEFILE_ZIP)

    assert len(records) == EXPECTED_HOTSPOT_COUNT
    assert [record.area_cd for record in records] == sorted(
        record.area_cd for record in records
    )
    assert {record.geometry_version for record in records} == {
        POLYGON_SHADOW_GEOMETRY_VERSION
    }
    assert all(record.geometry.is_valid for record in records)
    assert all(
        record.geometry.geom_type in {"Polygon", "MultiPolygon"}
        for record in records
    )
    assert {
        record.area_cd
        for record in records
        if record.normalization == "make_valid"
    } == {"POI070"}


def test_all_representative_points_are_wgs84_coordinates_inside_seoul() -> None:
    records = load_hotspot_master(XLSX, SHAPEFILE_ZIP)
    min_lng, min_lat, max_lng, max_lat = SEOUL_WGS84_BBOX

    assert all(min_lng <= record.lng <= max_lng for record in records)
    assert all(min_lat <= record.lat <= max_lat for record in records)

    with ZipFile(SHAPEFILE_ZIP) as zip_file:
        stem = next(
            name[:-4]
            for name in zip_file.namelist()
            if name.lower().endswith(".shp")
        )
        reader = shapefile.Reader(
            shp=BytesIO(zip_file.read(stem + ".shp")),
            shx=BytesIO(zip_file.read(stem + ".shx")),
            dbf=BytesIO(zip_file.read(stem + ".dbf")),
            encoding="utf-8",
        )
    geometries = {
        shape_record.record.as_dict()["AREA_CD"]: make_valid(
            shape(shape_record.shape.__geo_interface__)
        )
        for shape_record in reader.iterShapeRecords()
    }
    reader.close()
    assert all(
        Point(record.lng, record.lat).within(geometries[record.area_cd])
        for record in records
    )


def test_rejects_duplicate_area_code_in_xlsx(tmp_path: Path) -> None:
    workbook = _load_fixture_workbook()
    worksheet = workbook["장소목록"]
    worksheet.cell(row=3, column=3, value=worksheet.cell(row=2, column=3).value)
    duplicate_xlsx = tmp_path / "duplicate.xlsx"
    workbook.save(duplicate_xlsx)
    workbook.close()

    with pytest.raises(HotspotMasterError, match="duplicate AREA_CD in XLSX: POI001"):
        load_hotspot_master(duplicate_xlsx, SHAPEFILE_ZIP)


def test_rejects_missing_xlsx_place(tmp_path: Path) -> None:
    workbook = _load_fixture_workbook()
    worksheet = workbook["장소목록"]
    worksheet.delete_rows(2)
    incomplete_xlsx = tmp_path / "incomplete.xlsx"
    workbook.save(incomplete_xlsx)
    workbook.close()

    with pytest.raises(HotspotMasterError, match="exactly 121 hotspots; found 120"):
        load_hotspot_master(incomplete_xlsx, SHAPEFILE_ZIP)


def test_rejects_metadata_disagreement_between_sources(tmp_path: Path) -> None:
    workbook = _load_fixture_workbook()
    worksheet = workbook["장소목록"]
    worksheet.cell(row=2, column=4, value="다른 이름")
    changed_xlsx = tmp_path / "changed.xlsx"
    workbook.save(changed_xlsx)
    workbook.close()

    with pytest.raises(HotspotMasterError, match="metadata mismatch for POI001"):
        load_hotspot_master(changed_xlsx, SHAPEFILE_ZIP)


def test_rejects_duplicate_area_code_in_shapefile(tmp_path: Path) -> None:
    duplicate_zip = tmp_path / "duplicate.zip"
    _write_modified_shapefile_zip(duplicate_zip, duplicate_second_code=True)

    with pytest.raises(
        HotspotMasterError, match="duplicate AREA_CD in Shapefile: POI001"
    ):
        load_hotspot_master(XLSX, duplicate_zip)


def test_rejects_missing_shapefile_place(tmp_path: Path) -> None:
    incomplete_zip = tmp_path / "incomplete.zip"
    _write_modified_shapefile_zip(incomplete_zip, omit_last_record=True)

    with pytest.raises(HotspotMasterError, match="exactly 121 hotspots; found 120"):
        load_hotspot_master(XLSX, incomplete_zip)


def test_rejects_non_wgs84_shapefile(tmp_path: Path) -> None:
    projected_zip = tmp_path / "projected.zip"
    _write_modified_shapefile_zip(
        projected_zip,
        replacement_prj=b'PROJCS["Korea_2000_Korea_Central_Belt_2010"]',
    )

    with pytest.raises(HotspotMasterError, match="CRS must be WGS84"):
        load_hotspot_master(XLSX, projected_zip)
