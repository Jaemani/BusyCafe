"""Parse and join the official Seoul hotspot XLSX and area Shapefile.

The XLSX supplies the canonical place metadata.  The zipped Shapefile supplies
the corresponding area geometry.  Both sources are joined by ``AREA_CD`` and
must describe exactly the same 121 places; accepting a partial join would make
the resulting cafe scores silently incomplete.
"""

from __future__ import annotations

import warnings
from collections.abc import Iterable
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Final
from zipfile import BadZipFile, ZipFile

import shapefile
from shapely import make_valid
from openpyxl import load_workbook
from shapely.geometry import Point, shape
from shapely.geometry.base import BaseGeometry


EXPECTED_HOTSPOT_COUNT: Final = 121
MASTER_SHEET_NAME: Final = "장소목록"
REQUIRED_XLSX_COLUMNS: Final = ("CATEGORY", "AREA_CD", "AREA_NM")
REQUIRED_DBF_COLUMNS: Final = frozenset(REQUIRED_XLSX_COLUMNS)

# A data-integrity bound, not a scoring/tuning parameter.  It is deliberately
# broad enough to contain all of Seoul while detecting swapped/projected axes.
SEOUL_WGS84_BBOX: Final = (126.70, 37.40, 127.20, 37.72)


class HotspotMasterError(ValueError):
    """Raised when the official master files are incomplete or inconsistent."""


@dataclass(frozen=True, slots=True)
class HotspotMasterRecord:
    """A verified hotspot with a stable, geometry-derived representative point."""

    area_cd: str
    name: str
    category: str
    lat: float
    lng: float


@dataclass(frozen=True, slots=True)
class _MasterMetadata:
    area_cd: str
    name: str
    category: str


@dataclass(frozen=True, slots=True)
class _AreaRecord:
    metadata: _MasterMetadata
    geometry: BaseGeometry


def _required_text(value: object, *, field: str, row_number: int) -> str:
    if not isinstance(value, str) or not value.strip():
        raise HotspotMasterError(
            f"{field} must be non-empty text at XLSX row {row_number}"
        )
    return value.strip()


def _ensure_unique_codes(
    records: Iterable[_MasterMetadata], *, source: str
) -> dict[str, _MasterMetadata]:
    by_code: dict[str, _MasterMetadata] = {}
    duplicates: set[str] = set()
    for record in records:
        if record.area_cd in by_code:
            duplicates.add(record.area_cd)
        else:
            by_code[record.area_cd] = record
    if duplicates:
        joined = ", ".join(sorted(duplicates))
        raise HotspotMasterError(f"duplicate AREA_CD in {source}: {joined}")
    return by_code


def _read_xlsx(path: Path) -> dict[str, _MasterMetadata]:
    # openpyxl emits this while lazily iterating the official workbook, not
    # necessarily during load_workbook itself. Keep the suppression scoped to
    # the complete read and preserve every other warning.
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="^Unknown extension is not supported and will be removed$",
            category=UserWarning,
        )
        return _read_xlsx_contents(path)


def _read_xlsx_contents(path: Path) -> dict[str, _MasterMetadata]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (OSError, ValueError) as exc:
        raise HotspotMasterError(f"cannot read hotspot XLSX: {exc}") from exc

    try:
        if MASTER_SHEET_NAME not in workbook.sheetnames:
            raise HotspotMasterError(
                f"XLSX sheet {MASTER_SHEET_NAME!r} is missing"
            )
        worksheet = workbook[MASTER_SHEET_NAME]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            raise HotspotMasterError("hotspot XLSX is empty") from None

        header_indexes = {
            value.strip(): index
            for index, value in enumerate(header_row)
            if isinstance(value, str) and value.strip()
        }
        missing_columns = sorted(set(REQUIRED_XLSX_COLUMNS) - header_indexes.keys())
        if missing_columns:
            raise HotspotMasterError(
                "missing XLSX columns: " + ", ".join(missing_columns)
            )

        records: list[_MasterMetadata] = []
        for row_number, row in enumerate(rows, start=2):
            if not any(value is not None for value in row):
                continue
            records.append(
                _MasterMetadata(
                    area_cd=_required_text(
                        row[header_indexes["AREA_CD"]],
                        field="AREA_CD",
                        row_number=row_number,
                    ),
                    name=_required_text(
                        row[header_indexes["AREA_NM"]],
                        field="AREA_NM",
                        row_number=row_number,
                    ),
                    category=_required_text(
                        row[header_indexes["CATEGORY"]],
                        field="CATEGORY",
                        row_number=row_number,
                    ),
                )
            )
    finally:
        workbook.close()

    by_code = _ensure_unique_codes(records, source="XLSX")
    if len(by_code) != EXPECTED_HOTSPOT_COUNT:
        raise HotspotMasterError(
            f"XLSX must contain exactly {EXPECTED_HOTSPOT_COUNT} hotspots; "
            f"found {len(by_code)}"
        )
    return by_code


def _single_zip_member(zip_file: ZipFile, suffix: str) -> str:
    matches = sorted(
        name
        for name in zip_file.namelist()
        if not name.endswith("/") and name.lower().endswith(suffix)
    )
    if len(matches) != 1:
        raise HotspotMasterError(
            f"Shapefile ZIP must contain exactly one {suffix} file; "
            f"found {len(matches)}"
        )
    return matches[0]


def _validate_wgs84(prj: bytes) -> None:
    try:
        normalized = prj.decode("utf-8-sig").upper().replace(" ", "")
    except UnicodeDecodeError as exc:
        raise HotspotMasterError("Shapefile .prj is not UTF-8 text") from exc
    wgs84_markers = ("GCS_WGS_1984", "D_WGS_1984", 'SPHEROID["WGS_1984"')
    if not all(marker in normalized for marker in wgs84_markers):
        raise HotspotMasterError("Shapefile CRS must be WGS84 longitude/latitude")


def _validate_geometry(geometry: BaseGeometry, *, area_cd: str) -> Point:
    if geometry.geom_type not in {"Polygon", "MultiPolygon"}:
        raise HotspotMasterError(
            f"{area_cd} geometry must be Polygon or MultiPolygon"
        )
    if geometry.is_empty:
        raise HotspotMasterError(f"{area_cd} geometry is empty")
    # The verified official fixture contains one self-intersecting polygon
    # (POI070).  GEOS make_valid deterministically repairs it before deriving
    # the point, retaining all 121 official areas without trusting bad topology.
    if not geometry.is_valid:
        geometry = make_valid(geometry)
    if (
        geometry.geom_type not in {"Polygon", "MultiPolygon"}
        or geometry.is_empty
        or not geometry.is_valid
    ):
        raise HotspotMasterError(
            f"{area_cd} geometry could not be normalized to a valid polygon"
        )

    point = geometry.representative_point()
    if not point.within(geometry):
        raise HotspotMasterError(
            f"{area_cd} representative point is not inside its geometry"
        )
    min_lng, min_lat, max_lng, max_lat = SEOUL_WGS84_BBOX
    if not (min_lng <= point.x <= max_lng and min_lat <= point.y <= max_lat):
        raise HotspotMasterError(
            f"{area_cd} representative point is outside the Seoul WGS84 bbox"
        )
    return point


def _read_shapefile(path: Path) -> dict[str, _AreaRecord]:
    try:
        with ZipFile(path) as zip_file:
            members = {
                suffix: _single_zip_member(zip_file, suffix)
                for suffix in (".shp", ".shx", ".dbf", ".prj", ".cpg")
            }
            _validate_wgs84(zip_file.read(members[".prj"]))
            encoding = zip_file.read(members[".cpg"]).decode("ascii").strip()
            if encoding.upper().replace("-", "") != "UTF8":
                raise HotspotMasterError("Shapefile DBF encoding must be UTF-8")
            reader = shapefile.Reader(
                shp=BytesIO(zip_file.read(members[".shp"])),
                shx=BytesIO(zip_file.read(members[".shx"])),
                dbf=BytesIO(zip_file.read(members[".dbf"])),
                encoding="utf-8",
            )
    except (BadZipFile, OSError, UnicodeDecodeError, shapefile.ShapefileException) as exc:
        raise HotspotMasterError(f"cannot read hotspot Shapefile ZIP: {exc}") from exc

    dbf_columns = {field[0] for field in reader.fields[1:]}
    missing_columns = sorted(REQUIRED_DBF_COLUMNS - dbf_columns)
    if missing_columns:
        raise HotspotMasterError(
            "missing Shapefile DBF columns: " + ", ".join(missing_columns)
        )

    areas: list[_AreaRecord] = []
    try:
        for shape_record in reader.iterShapeRecords():
            attributes = shape_record.record.as_dict()
            metadata = _MasterMetadata(
                area_cd=str(attributes["AREA_CD"]).strip(),
                name=str(attributes["AREA_NM"]).strip(),
                category=str(attributes["CATEGORY"]).strip(),
            )
            if not all((metadata.area_cd, metadata.name, metadata.category)):
                raise HotspotMasterError(
                    "Shapefile AREA_CD, AREA_NM, and CATEGORY must be non-empty"
                )
            geometry = shape(shape_record.shape.__geo_interface__)
            _validate_geometry(geometry, area_cd=metadata.area_cd)
            areas.append(_AreaRecord(metadata=metadata, geometry=geometry))
    except (shapefile.ShapefileException, ValueError) as exc:
        if isinstance(exc, HotspotMasterError):
            raise
        raise HotspotMasterError(f"invalid Shapefile record: {exc}") from exc
    finally:
        reader.close()

    metadata_by_code = _ensure_unique_codes(
        (area.metadata for area in areas), source="Shapefile"
    )
    if len(metadata_by_code) != EXPECTED_HOTSPOT_COUNT:
        raise HotspotMasterError(
            f"Shapefile must contain exactly {EXPECTED_HOTSPOT_COUNT} hotspots; "
            f"found {len(metadata_by_code)}"
        )
    return {area.metadata.area_cd: area for area in areas}


def load_hotspot_master(
    xlsx_path: str | Path, shapefile_zip_path: str | Path
) -> tuple[HotspotMasterRecord, ...]:
    """Return the fully verified master, deterministically ordered by AREA_CD."""

    metadata_by_code = _read_xlsx(Path(xlsx_path))
    areas_by_code = _read_shapefile(Path(shapefile_zip_path))

    xlsx_codes = set(metadata_by_code)
    shape_codes = set(areas_by_code)
    if xlsx_codes != shape_codes:
        missing_geometry = ", ".join(sorted(xlsx_codes - shape_codes)) or "none"
        missing_metadata = ", ".join(sorted(shape_codes - xlsx_codes)) or "none"
        raise HotspotMasterError(
            "AREA_CD sets differ; "
            f"missing geometry: {missing_geometry}; "
            f"missing metadata: {missing_metadata}"
        )

    records: list[HotspotMasterRecord] = []
    for area_cd in sorted(xlsx_codes):
        metadata = metadata_by_code[area_cd]
        area = areas_by_code[area_cd]
        if metadata != area.metadata:
            raise HotspotMasterError(
                f"metadata mismatch for {area_cd}: XLSX={metadata!r}, "
                f"Shapefile={area.metadata!r}"
            )
        point = _validate_geometry(area.geometry, area_cd=area_cd)
        records.append(
            HotspotMasterRecord(
                area_cd=area_cd,
                name=metadata.name,
                category=metadata.category,
                lat=point.y,
                lng=point.x,
            )
        )
    return tuple(records)
